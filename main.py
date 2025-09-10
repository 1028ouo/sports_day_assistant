import speech_recognition as sr
import pyttsx3
import openpyxl
import os
from datetime import datetime
from gtts import gTTS
import pygame
import time
import pathlib

class SpeechToText:
  def __init__(self):
    self.rg = sr.Recognizer()

  def listen(self):
    with sr.Microphone() as source:
      self.rg.adjust_for_ambient_noise(source, duration=0.2)
      audioData = self.rg.listen(source)
      try:
        text = self.rg.recognize_google(audioData, language='zh-tw')
      except sr.UnknownValueError:
        text = ''
    return text

  def __call__(self):
    return self.listen()

# 將文字轉換成語音
def text_to_speech(command):
    # 初始化引擎
    engine = pyttsx3.init()
    engine.say(command)
    engine.runAndWait()

#獲取使用者的語音輸入
def get_user_input():
    text = ''
    while text == '':
        text = stt()
        if text == '':
            text_to_speech('聽不清楚，請再說一遍')
        else:
            print(text)
    return text

#語音辨識錯誤處理
def canNotRecognize(error_count, user_input, event):
  if error_count < 5:
    print(f"找不到'{event}':'{user_input}'，請確認環境是否無其他吵雜聲，並再重複一次'{event}' 。如要重複聽取問題，請說'再聽一次'，如要取消動作，請說'回上一步' 。")
    text_to_speech(f"找不到，{event}，{user_input}，請確認環境是否無其他吵雜聲，並再重複一次，{event} 。如要重複聽取問題，請說，再聽一次 。如要取消動作，請說，回上一步 。")
  elif error_count < 10:
    print(f"找不到'{event}':'{user_input}'，錯誤次數已達 {error_count} 次。請確認環境是否無其他吵雜聲，並再重複一次'{event}'。如要重複聽取問題，請說'再聽一次'，如要取消動作，請說'回上一步'。")
    text_to_speech(f"找不到，{event}，{user_input}，錯誤次數已達，{error_count}，次。請確認環境是否無其他吵雜聲，並再重複一次，{event} 。如要重複聽取問題，請說，再聽一次 。如要取消動作，請說，回上一步 。")
  else:
    print(f"找不到'{event}':'{user_input}'，錯誤次數已達 10 次，將自動返回上一步。")#(並且執行上一步的語音)(如果是一開始語音選擇的不會有這個)
    text_to_speech(f"找不到，{event}，{user_input}，錯誤次數已達，十，次。將自動返回上一步。")
  return

#找到歷年最佳的excel工作表
def load_workbook():
    script_dir = "運動會資料夾"  # 獲取程式碼檔案的路徑
    file_path = os.path.join(script_dir,'歷年最佳.xlsx')  # 組合成完整的檔案路徑
#     print(file_path)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)  # 使用完整的檔案路徑來載入 Excel 檔案
    else:
        wb = openpyxl.Workbook()  # 創建新的工作簿
        sheet = wb.active
        headers = ["賽事名", "賽事分類", "比賽日期", "選手資訊", "成績"]  # 欄位名稱
        sheet.append(headers)  # 將欄位名稱加入到工作表
        os.makedirs(script_dir, exist_ok=True)
        wb.save(file_path)  # 儲存工作簿
    return wb

wb= load_workbook()
sheet = wb.worksheets[0]

#找最佳成績
def find_event(event_name):
  # print("進入")
  # wb= load_workbook()
  # sheet = wb.worksheets[0]
  # print(sheet.max_row)
  for rowNum in range(2, sheet.max_row + 1):  # 從第二行開始讀取，因為第一行是標題
    # print(sheet.cell(rowNum, 1).value , event_name)
    if sheet.cell(rowNum, 1).value == event_name:  # 如果該行的賽事名稱與輸入的賽事名稱相同
      # print(sheet.cell(rowNum, 5).value,rowNum)
      return sheet.cell(rowNum, 5).value,rowNum
  return None,None

#將成績轉為數字
def convert_score(score_str):
  # 檢查成績字串中是否包含特定的字元
  if '公尺' in score_str:
    # 處理以公尺為單位的成績
    score, _ = score_str.split('公尺')
    score = float(score)
  elif '分' in score_str:
    # 處理以分鐘和秒為單位的成績
    minutes, seconds = score_str.split('分')
    seconds, _ = seconds.split('秒')
    score = float(minutes) * 60 + float(seconds)
  elif '秒' in score_str:
    # 處理以秒為單位的成績
    score, _ = score_str.split('秒')
    score = float(score)
  return score

#比較成績
def compare_scores(score_A, score_B):
  # 檢查成績的單位
  if '公尺' in score_A:
    # 將成績轉換為可以比較的數值
    score_A = convert_score(score_A)
    score_B = convert_score(score_B)
    # 對於以公尺為單位的成績，較大的數值是更好的成績
    return score_A > score_B
  else:
    # 將成績轉換為可以比較的數值
    score_A = convert_score(score_A)
    score_B = convert_score(score_B)
    # 對於以時間為單位的成績，較小的數值是更好的成績
    return score_A < score_B

# 呼叫函式，輸入你想要查詢的賽事名稱
def query_best():
  text_to_speech("請問你要查詢的賽事項目為?")
  user_input = get_user_input()
  score,row=find_event(user_input)
  if score==None:
    print("目前無該項目最佳成績紀錄")
  else:
    # print(score,row)
    # print("歷年最佳成績",sheet.cell(row, 1).value,  # 賽事名
    #   sheet.cell(row, 2).value,  # 賽事分類
    #   sheet.cell(row, 3).value,  # 比賽學年度
    #   sheet.cell(row, 4).value,  # 選手資訊
    #   sheet.cell(row, 5).value)  # 成績
    text=sheet.cell(row, 1).value+"的歷年最佳成績是："+score+"。紀錄保持者是："+str(sheet.cell(row, 3).value)+"學年度，"
    text+=sheet.cell(row, 1).value+"，"+sheet.cell(row, 2).value+"，"+sheet.cell(row, 4).value
    print(text)
    text_to_speech(text)

#創建資料夾
def create_event_folders(event_date, event_name):
  base_folder = "運動會資料夾"
  date_folder = os.path.join(base_folder, f"{event_date}")
  event_folder = os.path.join(date_folder, f"{event_name}")

  # 創建資料夾
  os.makedirs(event_folder, exist_ok=True)

  print(f"已成功創建資料夾結構：{event_folder}")

#創建excel檔(當沒找到符合需求的excel檔時)
def create_excel_file(excel_file_path):
  # 如果 Excel 檔案不存在，創建一個新檔案
  workbook = openpyxl.Workbook()

  # 選擇第一個工作表（假設只有一個工作表）
  sheet = workbook.active

  # 寫入標題行
  headers = ['選手資訊', '選手成績']
  sheet.append(headers)

  # 儲存 Excel 檔案
  workbook.save(excel_file_path)

#加入新資料進入excel檔
def add_data_to_excel(event_date, event_name, athlete_info, athlete_score, event_category):
#   script_dir = os.path.dirname(os.path.abspath(connection_file_path))  # 獲取程式碼檔案的路徑

  behind_path = f'運動會資料夾/{event_date}/{event_name}/{event_category}.xlsx'
#   excel_file_path = os.path.join(script_dir, behind_path)  # 組合成完整的檔案路徑
  excel_file_path=behind_path
  try:
    # 檢查 Excel 檔案是否存在，不存在則創建
    if not os.path.isfile(excel_file_path):
      create_excel_file(excel_file_path)

    # 打開現有的 Excel 檔案
    workbook = openpyxl.load_workbook(excel_file_path)

    # 選擇第一個工作表（假設只有一個工作表）
    sheet = workbook.active

    # 在最後一列新增資料
    row_data = [athlete_info, athlete_score]
    sheet.append(row_data)

    # 儲存 Excel 檔案
    workbook.save(excel_file_path)
  except Exception as e:
    print("尚未建立資料夾，請先新增賽事")
    text_to_speech("尚未建立資料夾，請先新增賽事")



#新增賽事
def create_event():
  # 提取新增賽事的相關資訊
  print('請問為第幾學年度?')
  text_to_speech('請問為第幾學年度?')
  user_input = get_user_input()

  event_date = user_input
  print('請問賽事名稱為?')
  text_to_speech('請問賽事名稱為?')
  user_input = get_user_input()
  event_name = user_input


  # 列印或進一步處理提取的資訊
  print("新增賽事:")
  print(f"賽事名: {event_name}")
  print(f"比賽學年度: {event_date}")
  text_to_speech("以上資料是否正確，如果正確，請回答'正確'，回到上一步，請說'回上一步'，如非以上話語，將自動跳轉回重新登記")
  user_input = get_user_input()
  if user_input=="回上一步":
    start()
  elif user_input=="正確":
    create_event_folders(event_date, event_name)
  else:
    create_event()


#登記成績
def record_score():
  text_to_speech('請問為第幾學年度?')
  user_input = get_user_input()
  event_date = user_input

  text_to_speech('請問賽事名稱為?')
  user_input = get_user_input()
  event_name = user_input

  text_to_speech('請問賽事分類為:第一項:初賽、第二項:複賽、第三項:決賽?請念第幾項做確認')
  user_input = get_user_input()
  while not any(event_category in user_input for event_category in ['第一項', '第二項', '第三項']):
    text_to_speech("無法分辨賽事分類，請再說一次")
    user_input = get_user_input()

  if user_input == '第一項':
    event_category = "初賽"
  elif user_input == '第二項':
    event_category = "複賽"
  elif user_input == '第三項':
    event_category = "決賽"
  text_to_speech("以上資料是否正確，如果正確，請回答'正確'，回到上一步，請說'回上一步'，如非以上話語，將自動跳轉回重新登記")

  user_input = get_user_input()
  if user_input=="回上一步":
    start()
  elif user_input!="正確":
    record_score()
    return
  else:
    count=0
    while True:
      text_to_speech('請問選手為?')
      user_input = get_user_input()
      athlete = user_input
      text_to_speech('請問選手成績為?')
      user_input = get_user_input()
      athlete_score = user_input
      if '點' in athlete_score:
        athlete_score = athlete_score.replace('點','.')

      print("登記現有賽事成績:")
      print(f"賽事名: {event_name}")
      print(f"選手資訊: {athlete}")
      print(f"選手成績: {athlete_score}")

      text_to_speech("以上資料是否正確，如果正確，請回答'正確'，回到上一步，請說'回上一步'，如非以上話語，將自動跳轉回重新登記")
      user_input = get_user_input()
      if user_input=="回上一步":
        record_score()
        return
      elif user_input!="正確":
        continue

      add_data_to_excel(event_date, event_name, athlete, athlete_score, event_category)

      if count==0:
        temp_event_name=event_name
        temp_event_category=event_category
        temp_event_date=event_date
        temp_athlete=athlete
        temp_score=athlete_score
        count+=1
      else:
        if compare_scores(athlete_score,temp_score):
          temp_score=athlete_score
          temp_athlete=athlete

      text_to_speech('請問要繼續登記嗎?請回答繼續登記或是停止登記')
      user_input = get_user_input()
      while user_input != "繼續登記" or user_input != "停止登記":
        text_to_speech('聽不清楚，請再說一遍')
        user_input = get_user_input()
      if user_input == "停止登記":
        #全部成績登記完後
        best_score,row=find_event(temp_event_name)
        if row==None:
          # print("新增")
          #創建欄位，加入temp_event_name,temp_event_category,temp_event_date,temp_athlete,temp_score
          sheet.append([temp_event_name, temp_event_category, temp_event_date, temp_athlete, temp_score])
        elif compare_scores(temp_score,best_score):
          # print("改變")
          #用temp_event_name,temp_event_category,temp_event_date,temp_athlete,temp_score更新"田徑 400公尺"這欄
          sheet.cell(row, 1).value = temp_event_name
          sheet.cell(row, 2).value = temp_event_category
          sheet.cell(row, 3).value = temp_event_date
          sheet.cell(row, 4).value = temp_athlete
          sheet.cell(row, 5).value = temp_score

#         best_script_dir = os.path.dirname(os.path.abspath(connection_file_path))  # 獲取程式碼檔案的路徑
        best_behind_path='運動會資料夾/歷年最佳.xlsx'
#         best_excel_file_path = os.path.join(best_script_dir, best_behind_path)  # 組合成完整的檔案路徑
        best_excel_file_path = best_behind_path
        wb.save(best_excel_file_path)  # 儲存變更
        break

def play_sound(file_path):
  pygame.mixer.init()
  pygame.mixer.music.load(file_path)
  pygame.mixer.music.play()
  while pygame.mixer.music.get_busy():
    pygame.time.Clock().tick(10)



# 把分轉換成秒
def time_to_seconds(time_str):
  if time_str is None:
    return 0  # 或者你可以根據需要返回其他值
  minutes, seconds = map(int, time_str.replace('分', ':').replace('秒', '').split(':'))
  return minutes * 60 + seconds

# 檢視
def view_records(event_name, event_category, event_date):
  # 為查詢生成Excel文件路徑
  excel_file_path = os.path.join('運動會資料夾', event_date, event_name, event_category, '賽事分類.xlsx')

  # 如果Excel文件存在，執行查詢
  if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    if sheet.cell(row=2, column=2).value[-1] == "秒":
      sorted_rows = sorted(sheet.iter_rows(min_row=2), key=lambda row: time_to_seconds(row[1].value))
    else:
      sorted_rows = sorted(sheet.iter_rows(min_row=2), key=lambda row: int(row[1].value.split('公尺')[0]), reverse=True)


        # 輸出
    print(f"收到！這邊已顯示{event_date} {event_name} {event_category} 的成績排名")

    for row in sorted_rows:
      athlete_info = row[0].value
      record = row[1].value
      print(f"| {athlete_info} | {record} |")

    workbook.close()
  else:
    #錯誤處理
    print("找不到相應的成績資料。將自動返回上一步")
    process_input("檢視成績")

#刪除
def delete_record(event_name, event_category, event_date, athlete_info):
  # 為查詢生成Excel文件路徑
  excel_file_path = os.path.join('運動會資料夾', event_date, event_name, event_category, '賽事分類.xlsx')

  # 如果Excel文件存在，執行刪除
  if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # 搜索並刪除相應的記錄
    for row in sheet.iter_rows(min_row=2):
      if row[0].value == athlete_info:
        sheet.delete_rows(row[0].row)
        break

    # 保存Excel文件
    workbook.save(excel_file_path)
    workbook.close()
    print(f"已刪除 {event_name}、{event_category}、{athlete_info} 的成績記錄。")
  else:
    print("找不到相應的成績資料。將自動返回上一步")
    process_input("刪除成績")

#宣布得獎名單
def announce_winners(event_name, event_category, event_date):
  # 為查詢生成賽事分類Excel文件路徑
  excel_file_path = os.path.join('運動會資料夾', event_date, event_name, event_category, '賽事分類.xlsx')

  # 如果賽事分類Excel文件存在，執行查詢並宣布得獎者
  if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # 進行排名
    if sheet.cell(row=2, column=2).value[-1] == "秒":
      sorted_rows = sorted(sheet.iter_rows(min_row=2), key=lambda row: time_to_seconds(row[1].value))
    else:
      sorted_rows = sorted(sheet.iter_rows(min_row=2), key=lambda row: int(row[1].value.split('公尺')[0]), reverse=True)

    # 準備文本轉語音輸出
    bgm_path = 'path_to_your_bgm.mp3'  # 替換為實際路徑
    applause_path = 'path_to_your_applause.mp3'  # 替換為實際路徑
    tts_texts = ['得獎名單生成中，請稍後！生成完畢，緊張刺激的一刻就要到來了！在三秒後開始宣布這次的得獎名單，三、二、一，開始宣布！']
    tts_texts.append('得獎名單：')

    # # 搜尋歷年最佳.xlsx文件
    # best_records_file_path = os.path.join('運動會資料夾', '歷年最佳.xlsx')

    # # 如果歷年最佳.xlsx文件存在，進行比較
    # if os.path.exists(best_records_file_path):
    #   best_records_workbook = openpyxl.load_workbook(best_records_file_path)
    #   best_records_sheet = best_records_workbook.active

    #   for row in best_records_sheet.iter_rows(min_row=2):
    #     athlete_info = row[0].value
    #     best_record = row[4].value  # Assuming the record is in the fifth column

    #     # 搜尋當前賽事分類.xlsx文件，找到對應選手的成績
    #     current_record = None
    #     for current_row in sorted_rows:
    #       if current_row[0].value == athlete_info:
    #         current_record = current_row[1].value
    #         break

    #     # 如果找到了當前賽事分類.xlsx中的選手成績，進行比較
    #     if current_record is not None and current_record < best_record:
    #       tts_texts.append(f"恭喜{athlete_info}突破大會紀錄！")

    #   best_records_workbook.close()


        # 宣布得獎者
    for idx, row in enumerate(sorted_rows, start=1):
      athlete_info = row[0].value
      record = row[1].value
      place = "第" + str(idx) + "名"

      print(f"恭喜 {athlete_info} 榮獲 {event_name}{place}！成績 {record}。")
      tts_texts.append(f"恭喜 {athlete_info} 榮獲 {event_name}{place}！成績 {record}。")

      #如果是最佳成績
      best_record,_=find_event(event_name)
      if best_record==record:
        tts_texts.append(f"恭喜{athlete_info}突破大會紀錄！")

      if idx == 3:
        break

    tts_texts.append('讓我們恭喜以上得獎者！')
    # 合併文本轉語音文本
    tts_text = ' '.join(tts_texts)

    # 創建gTTS對象，指定語言為中文 ('zh-tw' 或 'zh-cn')
    tts = gTTS(tts_text, lang='zh-cn')

    # 保存音頻文件
    tts.save('announcement.mp3')

    # 播放文本轉語音音頻
    play_sound('announcement.mp3')

    workbook.close()
  else:
    print("找不到相應的成績資料。將自動返回上一步")
    process_input("宣布得獎名單")

#辨別與確認執行操作
def process_input(user_input):
  if "新增賽事" in user_input:
    create_event()

  elif "登記成績" in user_input:
    record_score()

  elif "檢視成績" in user_input:
    text_to_speech('請說出賽事學年度:')
    user_input = get_user_input()
    event_date_input = user_input

    text_to_speech('請說出賽事名稱:')
    user_input = get_user_input()
    event_name = user_input

    # text_to_speech('請說出賽事分類:')
    # user_input = get_user_input()
    # event_category = user_input
    text_to_speech('請問賽事分類為:第一項:初賽、第二項:複賽、第三項:決賽?請念第幾項做確認')
    user_input = get_user_input()
    while not any(event_category in user_input for event_category in ['第一項', '第二項', '第三項']):
      text_to_speech("無法分辨賽事分類，請再說一次")
      user_input = get_user_input()

    if user_input == '第一項':
      event_category = "初賽"
    elif user_input == '第二項':
      event_category = "複賽"
    elif user_input == '第三項':
      event_category = "決賽"

    view_records(event_name, event_category, event_date_input)

  elif "刪除成績" in user_input:
    text_to_speech('請說出賽事學年度:')
    user_input = get_user_input()
    event_date_input = user_input

    text_to_speech('請說出賽事名稱:')
    user_input = get_user_input()
    event_name = user_input

    text_to_speech('請問賽事分類為:第一項:初賽、第二項:複賽、第三項:決賽?請念第幾項做確認')
    user_input = get_user_input()
    while not any(event_category in user_input for event_category in ['第一項', '第二項', '第三項']):
      text_to_speech("無法分辨賽事分類，請再說一次")
      user_input = get_user_input()

    if user_input == '第一項':
      event_category = "初賽"
    elif user_input == '第二項':
      event_category = "複賽"
    elif user_input == '第三項':
      event_category = "決賽"

    text_to_speech('請說出欲刪除的選手名稱:')
    user_input = get_user_input()
    athlete_info = user_input

    delete_record(event_name, event_category, event_date_input, athlete_info)

  elif "宣布得獎名單" in user_input:
    text_to_speech('請說出賽事學年度:')
    user_input = get_user_input()
    event_date_input = user_input

    text_to_speech('請說出賽事名稱:')
    user_input = get_user_input()
    event_name = user_input

    text_to_speech('請問賽事分類為:第一項:初賽、第二項:複賽、第三項:決賽?請念第幾項做確認')
    user_input = get_user_input()
    while not any(event_category in user_input for event_category in ['第一項', '第二項', '第三項']):
      text_to_speech("無法分辨賽事分類，請再說一次")
      user_input = get_user_input()

    if user_input == '第一項':
      event_category = "初賽"
    elif user_input == '第二項':
      event_category = "複賽"
    elif user_input == '第三項':
      event_category = "決賽"

    # announce_winners(event_name, event_category, event_date_input)
  elif "查詢歷年最佳" in user_input:
    query_best()
  else:
    print("無法識別的指令")
    start()

def start():
  print('請問您想執行的操作為？選項：新增賽事、登記成績、檢視成績、刪除成績、宣布得獎名單、查詢歷年最佳')
  text_to_speech('請問您想執行的操作為？選項：新增賽事、登記成績、檢視成績、刪除成績、宣布得獎名單、查詢歷年最佳')
  user_input = get_user_input()

  text_to_speech(user_input)
  process_input(user_input)

if __name__ == "__main__":
  stt = SpeechToText()  # 創建 SpeechToText 實例
  start()